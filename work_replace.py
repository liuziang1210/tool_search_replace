import os
import sys
os.environ["GREENCZ"] = "C:/SystemAnalyst/SE_PROD/scriptsResources"  #
sys.path.append(os.path.expandvars("$GREENCZ/openpyxl-3.0.3"))
sys.path.append(os.path.expandvars("$GREENCZ/jdcal-1.4.1"))
sys.path.append(os.path.expandvars("$GREENCZ/et_xmlfile-1.0.1"))
import time
from typing import List, Dict
import multiprocessing
#from multiprocessing import Pool
from multiprocessing.dummy import Pool
import tkinter as tk
from tkinter import filedialog, ttk
import webbrowser
import openpyxl
from openpyxl import utils
from xlrd3 import open_workbook
import re

SEARCH_RESULT_DICT = {}

def get_the_file_name_num(folder_path: str):
    xml_file_names = {}
    count1 = 0
    count2 = 0
    for filename in os.listdir(folder_path):
        if not filename.startswith('~$'):
            if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
                file_path = os.path.join(folder_path, filename)
                file_name = os.path.splitext(filename)[0]
                file_path = file_path.replace('\\', '/')
                xml_file_names[file_name] = file_path
                count1 += 1
            if filename.endswith('.xls'):
                file_path = os.path.join(folder_path, filename)
                file_name = os.path.splitext(filename)[0]
                file_path = file_path.replace('\\', '/')
                xml_file_names[file_name + ".xls"] = file_path
                count2 += 1
    count = count1 + count2
    return xml_file_names, count, count1, count2

def search_word_in_xlsx_file(file_path: str, search_word: str):
    word_found = False
    count_for_xlsx = 0
    results = []
    search_word_lower = search_word.lower()
    if file_path.endswith('.xlsm') or file_path.endswith('.xlsx'):
        count_for_xlsx += 1
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_index, col in enumerate(row, start=1):
                    if col and search_word_lower in str(col).lower():
                        cell_coordinate = sheet.cell(row=row_index, column=col_index).coordinate
                        results.append((file_path, sheet_name, cell_coordinate))
                        word_found = True
    total_num = count_for_xlsx
    return results, total_num, word_found

def search_word_in_xls_files(excel_files: Dict, search_word: str, output_text):
    word_found = False
    count_for_xls = 0
    start_time = time.time()
    for file_name, file_path in excel_files.items():
        if file_path.endswith('.xls'):
            count_for_xls += 1
            workbook_xls = open_workbook(file_path)
            for sheet_name in workbook_xls.sheet_names():
                sheet = workbook_xls.sheet_by_name(sheet_name)
                for row_index in range(sheet.nrows):
                    for col_index in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_index, col_index)
                        if cell_value:
                            if search_word.lower() in str(cell_value).lower():
                                cell_row = row_index + 1
                                cell_col = utils.get_column_letter(col_index + 1)
                                cell_coordinate = f"{cell_col}{cell_row}"
                                description_text = f"The Words \"{search_word}\" found in "
                                output_text.insert(tk.END, description_text)
                                start_index = output_text.index(tk.END)
                                unique_tag = f"link_{start_index.replace('.', '_')}"
                                output_text.insert(tk.END, file_name, unique_tag)
                                output_text.tag_add(unique_tag, start_index, f"{start_index} + {len(file_name)}c")
                                remaining_text = f", sheet: {sheet_name}, cell: {cell_coordinate}\n"
                                output_text.insert(tk.END, remaining_text)
                                output_text.tag_bind(unique_tag, "<Button-1>",lambda event, path=file_path: open_file(path))
                                output_text.tag_config(unique_tag, foreground="blue", underline=1)
                                word_found = True
                                output_text.update_idletasks()
    end_time = time.time()
    total_time = end_time - start_time
    total_num = count_for_xls
    if not word_found:
        output_text.insert(tk.END, f"The word '{search_word}' was not found in any of the .xls files.\n")
    output_text.insert(tk.END, f"Total .xls files processed: {total_num}\n")
    output_text.insert(tk.END, f"Total time taken: {total_time:.2f} seconds\n")


def replace_word_in_xlsx_file(file_path: str, search_word: str, replace_word: str, search_result):
    word_found = False
    count_for_xlsx = 0
    count_for_replace_cell = 0
    if file_path.endswith('.xlsx') or file_path.endswith('.xlsm'):
        count_for_xlsx += 1
        workbook = openpyxl.load_workbook(file_path)
        pattern = re.compile(re.escape(search_word), re.IGNORECASE)
        for sheet_name, cell_coordinate in search_result:
            sheet = workbook[sheet_name]
            row, col = openpyxl.utils.coordinate_to_tuple(cell_coordinate)
            cell = sheet.cell(row=row, column=col)
            if cell.value and search_word.lower() in str(cell.value).lower():

                cell.value = pattern.sub(replace_word, str(cell.value))
                count_for_replace_cell +=1
                word_found = True
        workbook.save(file_path)
        total_num = count_for_xlsx
        return [], total_num, word_found, count_for_replace_cell

def open_file(file_path):
    webbrowser.open(file_path)

def main():
    root = tk.Tk()
    root.title("Excel File Search and Replace")

    tab_control = ttk.Notebook(root)

    search_tab = ttk.Frame(tab_control)
    tab_control.add(search_tab, text="Search")

    replace_tab = ttk.Frame(tab_control)
    tab_control.add(replace_tab, text="Replace")

    tab_control.pack(expand=1, fill="both")

    def select_folder(entry_field):
        folder_path = filedialog.askdirectory()
        entry_field.delete(0, tk.END)
        entry_field.insert(0, folder_path)

    def reset_fields(entry_fields: List[tk.Entry], text_widget: tk.Text):
        for entry in entry_fields:
            entry.delete(0, tk.END)
        text_widget.delete("1.0", tk.END)

        for widget in replace_file_frame.winfo_children():
            if isinstance(widget, tk.Checkbutton):
                widget.deselect()

    def start_search():
        folder_path = folder_entry.get()
        search_word = search_entry.get()

        SEARCH_RESULT_DICT.clear()

        if not search_word.strip():
            output_text.insert(tk.END, "Please enter a search word.\n")
            return

        xml_files, count, count1, count2 = get_the_file_name_num(folder_path)

        output_text.delete("1.0", tk.END)
        output_text.insert(tk.END, f'In folder {folder_path}, a total of {count} files were found: {count1} .xlsx/.xlsm files and {count2} .xls files.\n')
        output_text.update_idletasks()
        output_text.insert(tk.END, f'--------------------------------------------------\n')
        output_text.update_idletasks()

        if count1 != 0:
            output_text.insert(tk.END, f'start process .xlsx/.xlsm...\n')
            output_text.update_idletasks()
            start_time = time.time()
            xlsx_files = [file_path for file_path in xml_files.values() if file_path.endswith('.xlsx') or file_path.endswith('.xlsm')]

            pool = Pool(processes=multiprocessing.cpu_count())
            results = pool.starmap(search_word_in_xlsx_file, [(file, search_word) for file in xlsx_files])

            pool.close()
            pool.join()

            end_time = time.time()
            total_time = end_time - start_time

            total_xlsx_num = 0
            total_word_found = False
            all_results = []

            for res in results:
                result, xlsx_num, word_found = res
                total_xlsx_num += xlsx_num
                total_word_found = total_word_found or word_found
                all_results.extend(result)

            for (file_path, sheet_name, cell_coordinate) in all_results:
                if file_path not in SEARCH_RESULT_DICT:
                    SEARCH_RESULT_DICT[file_path] = []
                SEARCH_RESULT_DICT[file_path].append((sheet_name, cell_coordinate))

            for (file_path, sheet_name, cell_coordinate) in all_results:
                file_name = os.path.basename(file_path)
                description_text = f"The Words \"{search_word}\" found in "
                output_text.insert(tk.END, description_text)
                start_index = output_text.index(tk.END)
                unique_tag = f"link_{start_index.replace('.', '_')}"
                output_text.insert(tk.END, file_name, unique_tag)
                output_text.tag_add(unique_tag, start_index, f"{start_index} + {len(file_name)}c")
                remaining_text = f", sheet: {sheet_name}, cell: {cell_coordinate}\n"
                output_text.insert(tk.END, remaining_text)
                output_text.tag_bind(unique_tag, "<Button-1>", lambda event, path=file_path: open_file(path))
                output_text.tag_config(unique_tag, foreground="blue", underline=True)

            if not total_word_found:
                output_text.insert(tk.END, f"The word '{search_word}' was not found in any of the .xlsx/.xlsm files.\n")
            output_text.insert(tk.END, f"Total .xlsx/.xlsm files processed: {total_xlsx_num}\n")
            output_text.insert(tk.END, f"Total time taken: {total_time:.2f} seconds\n")
            output_text.insert(tk.END, f'end process .xlsx/.xlsm\n')
            output_text.update_idletasks()
        else:
            output_text.insert(tk.END, f'No .xlsm/.xlsx files\n')

        if count2 != 0:
            output_text.insert(tk.END, f'--------------------------------------------------\n')
            output_text.insert(tk.END, f'start process .xls...\n')
            search_word_in_xls_files(xml_files, search_word, output_text)
            output_text.insert(tk.END, f'end process .xls\n')
        else:
            output_text.insert(tk.END, f'--------------------------------------------------\n')
            output_text.insert(tk.END, f'No .xls files\n')

        output_text.insert(tk.END, f'--------------------------------------------------\n')
        output_text.insert(tk.END, f'End! Thank you for using this program!')


        replace_search_entry.config(state='normal')
        replace_search_entry.delete(0, tk.END)
        replace_search_entry.insert(0, search_word)
        replace_search_entry.config(state='readonly')

        for widget in replace_file_frame.winfo_children():
            widget.destroy()

        for result in all_results:
            file_path, sheet_name, cell_coordinate = result
            file_name = os.path.basename(file_path)
            display_text = f"{file_name} - {sheet_name} - {cell_coordinate}"
            checkbox_var = tk.BooleanVar(value=True)
            checkbox = tk.Checkbutton(replace_file_frame, text=display_text, variable=checkbox_var)
            checkbox.var = checkbox_var
            checkbox.file_path = file_path
            checkbox.pack(anchor='w')

    def start_replace():
        replace_word = replace_entry.get()
        search_word = replace_search_entry.get()

        if not replace_word.strip() or not search_word.strip():
            replace_output_text.insert(tk.END, "Please enter both search and replace words.\n")
            return

        replace_output_text.delete("1.0", tk.END)
        replace_output_text.insert(tk.END, f'Starting replace for the word "{search_word}" with "{replace_word}".\n')
        replace_output_text.update_idletasks()
        replace_output_text.insert(tk.END, f'--------------------------------------------------\n')
        replace_output_text.update_idletasks()

        selected_files_and_cells = {}

        for widget in replace_file_frame.winfo_children():
            if isinstance(widget, tk.Checkbutton) and widget.var.get():
                file_path = widget.file_path
                sheet_name, cell_coordinate = widget.cget("text").split(" - ")[1:]
                if file_path not in selected_files_and_cells:
                    selected_files_and_cells[file_path] = []
                selected_files_and_cells[file_path].append((sheet_name, cell_coordinate))
                print(f"the file is {selected_files_and_cells}")

        if selected_files_and_cells:
            replace_output_text.insert(tk.END, f'start process .xlsx/.xlsm...\n')
            replace_output_text.update_idletasks()
            start_time = time.time()

            pool = Pool(processes=multiprocessing.cpu_count())
            try:
                results = pool.starmap(replace_word_in_xlsx_file,
                                       [(file_path, search_word, replace_word, search_results) for
                                        file_path, search_results
                                        in selected_files_and_cells.items()])
            except Exception as e:
                replace_output_text.insert(tk.END, f"ERROR: {e}\n")
                return str(e)
            pool.close()
            pool.join()

            end_time = time.time()
            total_time = end_time - start_time

            total_xlsx_num = 0
            total_cells_replaced = 0
            total_word_found = False

            for res in results:
                result, xlsx_num, word_found, count_for_replace_cell  = res
                total_xlsx_num += xlsx_num
                total_word_found = total_word_found or word_found
                total_cells_replaced += count_for_replace_cell

            if not total_word_found:
                replace_output_text.insert(tk.END,
                                           f"The word '{search_word}' was not found in any of the .xlsx/.xlsm files.\n")
            replace_output_text.insert(tk.END, f"Total .xlsx/.xlsm files processed: {total_xlsx_num}\n")
            replace_output_text.insert(tk.END, f"Total cells replaced:: {total_cells_replaced}\n")
            replace_output_text.insert(tk.END, f"Total time taken: {total_time:.2f} seconds\n")
            replace_output_text.insert(tk.END, f'end process .xlsx/.xlsm\n')
            replace_output_text.update_idletasks()
        else:
            replace_output_text.insert(tk.END, f'No files selected for replacement.\n')



        replace_output_text.insert(tk.END, f'--------------------------------------------------\n')
        replace_output_text.insert(tk.END, f'End! Thank you for using this program!')


    # Search Tab Widgets
    folder_label = tk.Label(search_tab, text="Folder Path:")
    folder_label.grid(row=0, column=0, padx=10, pady=10, sticky='w')

    folder_entry = tk.Entry(search_tab, width=50)
    folder_entry.grid(row=0, column=1, padx=10, pady=10)

    folder_button = tk.Button(search_tab, text="Select Folder", command=lambda: select_folder(folder_entry))
    folder_button.grid(row=0, column=2, padx=10, pady=10)

    search_label = tk.Label(search_tab, text="Search Word:")
    search_label.grid(row=1, column=0, padx=10, pady=10)

    search_entry = tk.Entry(search_tab, width=50)
    search_entry.grid(row=1, column=1, padx=10, pady=10)

    start_button = tk.Button(search_tab, text="Start Search", command=start_search)
    start_button.grid(row=2, column=1, padx=10, pady=10)

    reset_button = tk.Button(search_tab, text="Reset", command=lambda: reset_fields([folder_entry, search_entry], output_text))
    reset_button.grid(row=2, column=2, padx=10, pady=10)

    output_text = tk.Text(search_tab, height=10, width=80)
    output_text.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky='nsew')

    # Replace Tab Widgets
    replace_search_label = tk.Label(replace_tab, text="Search Word:")
    replace_search_label.grid(row=0, column=0, padx=10, pady=10)

    replace_search_entry = tk.Entry(replace_tab, width=50, state="readonly")
    replace_search_entry.grid(row=0, column=1, padx=10, pady=10)

    replace_label = tk.Label(replace_tab, text="Replace With:")
    replace_label.grid(row=1, column=0, padx=10, pady=10)

    replace_entry = tk.Entry(replace_tab, width=50)
    replace_entry.grid(row=1, column=1, padx=10, pady=10)

    replace_start_button = tk.Button(replace_tab, text="Start Replace", command=start_replace)
    replace_start_button.grid(row=2, column=1, padx=10, pady=10)

    replace_reset_button = tk.Button(replace_tab, text="Reset", command=lambda: reset_fields([replace_search_entry, replace_entry], replace_output_text))
    replace_reset_button.grid(row=2, column=2, padx=10, pady=10)

########   tk.Listbox   ###############
    # replace_file_listbox = tk.Listbox(replace_tab, selectmode="multiple", width=80, height=10)
    # replace_file_listbox.grid(row=3, column=0, columnspan=3, padx=10, pady=10)

##########################################

    replace_canvas = tk.Canvas(replace_tab)
    replace_canvas.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

    replace_scrollbar = tk.Scrollbar(replace_tab, orient="vertical", command=replace_canvas.yview)
    replace_scrollbar.grid(row=3, column=3, sticky="ns")

    replace_canvas.configure(yscrollcommand=replace_scrollbar.set)

    replace_file_frame = tk.Frame(replace_canvas)
    replace_canvas.create_window((0, 0), window=replace_file_frame, anchor="nw")

    replace_file_frame.bind("<Configure>", lambda e: replace_canvas.configure(scrollregion=replace_canvas.bbox("all")))

    replace_output_text = tk.Text(replace_tab, height=10, width=80)
    replace_output_text.grid(row=4, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)
    search_tab.grid_rowconfigure(3, weight=1)
    search_tab.grid_columnconfigure(1, weight=1)
    replace_tab.grid_rowconfigure(3, weight=1)
    replace_tab.grid_columnconfigure(1, weight=1)

    root.mainloop()

if __name__ == "__main__":
    main()

############ to test send
####second
#####3333
########4444444